"""``repositories.xlsx`` bulk input (spec section 50).

Columns: ``Repository URL``, ``Branch``, ``Analysis Mode``, ``Priority``. Each valid row
upserts a ``Repository`` and enqueues an ordinary ``AnalysisRun`` through the same
``JobManager`` path as ``POST /repositories/{id}/runs`` - no separate enqueue engine.
"""

from __future__ import annotations

import hashlib
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from archon.core.errors import ArchonError, validation
from archon.core.ids import new_id
from archon.core.logging import get_logger
from archon.db.models import Repository
from archon.domain.enums import RunMode
from archon.jobs.manager import JobManager
from archon.providers.repo import provider_for

log = get_logger("archon.reporting.bulk")

_COLUMNS = ("repository url", "branch", "analysis mode", "priority")
_jobs = JobManager()


@dataclass
class BulkRowResult:
    row: int
    url: str
    status: Literal["created", "skipped", "error"]
    reason: str = ""
    run_id: str | None = None

    def as_dict(self) -> dict:
        return {
            "row": self.row, "url": self.url, "status": self.status,
            "reason": self.reason, "run_id": self.run_id,
        }


def _norm_header(value: object) -> str:
    return str(value or "").strip().lower()


def _config_hash(mode: str, ref: str | None) -> str:
    return hashlib.sha256(f"{mode}|{ref or ''}".encode()).hexdigest()[:32]


def import_repositories_xlsx(session: Session, src: str | Path | bytes) -> list[BulkRowResult]:
    data = src if isinstance(src, bytes) else Path(src).read_bytes()
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise validation("repositories.xlsx is empty")
    header = [_norm_header(c) for c in rows[0]]
    missing = [c for c in _COLUMNS[:3] if c not in header]
    if missing:
        raise validation(
            f"repositories.xlsx missing required column(s): {', '.join(missing)}"
        )
    idx = {name: header.index(name) for name in _COLUMNS if name in header}

    def _get(raw: tuple, name: str) -> str | None:
        i = idx.get(name)
        if i is None or i >= len(raw):
            return None
        v = raw[i]
        return str(v).strip() if v not in (None, "") else None

    results: list[BulkRowResult] = []
    for n, raw in enumerate(rows[1:], start=2):
        url = _get(raw, "repository url")
        if not url:
            continue  # blank line
        branch = _get(raw, "branch")
        mode_s = (_get(raw, "analysis mode") or "ANALYSIS_ONLY").upper()
        priority_s = _get(raw, "priority")

        try:
            mode = RunMode(mode_s)
        except ValueError:
            results.append(BulkRowResult(n, url, "error", f"invalid Analysis Mode {mode_s!r}"))
            continue
        try:
            priority = int(priority_s) if priority_s is not None else 100
        except ValueError:
            results.append(BulkRowResult(n, url, "error", f"invalid Priority {priority_s!r}"))
            continue

        try:
            provider = provider_for(url)
            ref = provider.parse(url)
        except ArchonError as exc:
            results.append(BulkRowResult(n, url, "error", exc.message))
            continue

        repo = session.scalar(
            select(Repository).where(
                Repository.provider == provider.kind, Repository.url == ref.canonical_url
            )
        )
        if repo is None:
            repo = Repository(
                id=new_id("repo"), provider=provider.kind, url=ref.canonical_url,
                owner=ref.owner, name=ref.name,
            )
            session.add(repo)
            session.flush()

        try:
            job = _jobs.create_run_with_job(
                session, repository_id=repo.id, mode=mode, requested_ref=branch,
                config_hash=_config_hash(mode.value, branch), priority=priority,
            )
        except ArchonError as exc:
            results.append(BulkRowResult(n, url, "skipped", exc.message))
            continue

        results.append(BulkRowResult(n, url, "created", run_id=job.run_id))

    log.info(
        "bulk import complete",
        extra={"extra_fields": {
            "rows": len(results),
            "created": sum(1 for r in results if r.status == "created"),
            "skipped": sum(1 for r in results if r.status == "skipped"),
            "errors": sum(1 for r in results if r.status == "error"),
        }},
    )
    return results

"""Render ``ARCHON_Legacy_Analysis.xlsx`` (spec sections 49-50).

14 sheets, one per analysis domain, each sourced from :mod:`archon.reporting.queries`
(the same code path as the JSON API). Pure presentation - no engine, no new queries.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from archon.reporting import queries

REPORT_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
REPORT_FILENAME = "ARCHON_Legacy_Analysis.xlsx"

_HEADER_FONT = Font(bold=True)


def _cell(value: Any) -> Any:
    """Coerce a value openpyxl cannot store natively into a string."""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _write_table(ws: Worksheet, headers: list[str], rows: list[dict]) -> None:
    ws.append(headers)
    for c in ws[1]:
        c.font = _HEADER_FONT
    for row in rows:
        ws.append([_cell(row.get(h)) for h in headers])
    for i, h in enumerate(headers, start=1):
        width = max(len(h), *(len(str(r.get(h, ""))) for r in rows)) if rows else len(h)
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(width + 2, 60)


def _kv(ws: Worksheet, pairs: list[tuple[str, Any]]) -> None:
    ws.append(["Field", "Value"])
    for c in ws[1]:
        c.font = _HEADER_FONT
    for k, v in pairs:
        ws.append([k, _cell(v)])
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 70


def _dump(rows: list) -> list[dict]:
    return [r.model_dump() if hasattr(r, "model_dump") else dict(r) for r in rows]


def _headers(rows: list[dict], preferred: list[str]) -> list[str]:
    if not rows:
        return preferred
    keys = list(rows[0].keys())
    ordered = [k for k in preferred if k in keys]
    return ordered + [k for k in keys if k not in ordered]


# --- sheets ---------------------------------------------------------------------------


def _sheet_executive_summary(wb: Workbook, session: Session, run_id: str) -> None:
    ws = wb.create_sheet("Executive Summary")
    run = queries.run_overview(session, run_id)
    ld = queries.legacy_dna(session, run_id)
    hs = queries.hotspots(session, run_id)
    td = queries.technical_debt(session, run_id)
    cs = queries.change_safety(session, run_id)
    fails = queries.failures(session, run_id)
    patches = queries.patches(session, run_id)
    incs = queries.incidents_for_run(session, run_id)
    mod = queries.modernization(session, run_id)
    und = queries.understanding(session, run_id)

    def _cat_count(rows, attr, value):
        return sum(1 for r in rows if getattr(r, attr, None) == value)

    _kv(ws, [
        ("Run id", run_id),
        ("Repository id", run.repository_id),
        ("Commit", run.snapshot.commit_sha if run.snapshot else None),
        ("State", run.state),
        ("Last completed stage", run.last_completed_stage),
        ("Mode", run.mode),
        ("Started", run.started_at),
        ("Ended", run.ended_at),
        ("Repository understanding score", getattr(und, "overall_score", None)),
        ("Understanding confidence", getattr(und, "confidence", None)),
        ("Components scored (Legacy DNA)", len(ld)),
        ("  CRITICAL / HIGH legacy risk", _cat_count(ld, "category", "CRITICAL") + _cat_count(ld, "category", "HIGH")),
        ("Hotspots (RISKY / CRITICAL)", _cat_count(hs, "classification", "RISKY") + _cat_count(hs, "classification", "CRITICAL")),
        ("Technical-debt findings", len(td)),
        ("  HIGH / CRITICAL severity", _cat_count(td, "severity", "HIGH") + _cat_count(td, "severity", "CRITICAL")),
        ("Change-safety: RISKY / DANGEROUS", _cat_count(cs, "risk_category", "RISKY") + _cat_count(cs, "risk_category", "DANGEROUS")),
        ("Test failures", len(fails)),
        ("Candidate patches", len(patches)),
        ("  VERIFIED", _cat_count(patches, "state", "VERIFIED")),
        ("Incidents recorded", len(incs)),
        ("Modernization recommendations", len(mod)),
        ("Evidence rows", len(queries.evidence(session, run_id))),
    ])


def _sheet_repository_understanding(wb: Workbook, session: Session, run_id: str) -> None:
    ws = wb.create_sheet("Repository Understanding")
    und = queries.understanding(session, run_id)
    if und is None:
        _kv(ws, [("status", "not scored for this run")])
        return
    _kv(ws, [("Overall score", und.overall_score), ("Confidence", und.confidence)])
    ws.append([])
    ws.append(["Dimension", "Score"])
    for c in ws[ws.max_row]:
        c.font = _HEADER_FONT
    for d in und.dimensions:
        ws.append([d.name, d.score])
    ws.append([])
    ws.append(["Evidence coverage", ""])
    ws[ws.max_row][0].font = _HEADER_FONT
    for k, v in (und.evidence_coverage or {}).items():
        ws.append([k, _cell(v)])


def _sheet_architecture(wb: Workbook, session: Session, run_id: str) -> None:
    ws = wb.create_sheet("Architecture")
    arch = queries.architecture_report(session, run_id)
    if arch is None:
        _kv(ws, [("status", "architecture not reconstructed for this run")])
        return
    rows = _dump(arch.modules)
    _write_table(ws, _headers(rows, [
        "qualified_name", "role", "fan_in", "fan_out", "instability",
        "betweenness_centrality", "in_cycle", "scc_size", "path",
    ]), rows)
    ws.append([])
    ws.append(["Import cycles", "; ".join(" -> ".join(c) for c in (arch.cycles or [])) or "none"])
    ws.append(["Roles", "; ".join(f"{k}={v}" for k, v in (arch.roles or {}).items())])


def _sheet_legacy_dna(wb: Workbook, session: Session, run_id: str) -> None:
    ws = wb.create_sheet("Legacy DNA")
    rows = _dump(queries.legacy_dna(session, run_id))
    _write_table(ws, _headers(rows, [
        "component_qn", "legacy_risk_score", "category", "confidence", "complexity",
        "churn", "coupling", "coverage", "coverage_is_proxy", "debt_score",
        "assumption_count", "age_days",
    ]), rows)


def _sheet_change_safety(wb: Workbook, session: Session, run_id: str) -> None:
    ws = wb.create_sheet("Change Safety")
    rows = _dump(queries.change_safety(session, run_id))
    _write_table(ws, _headers(rows, [
        "component_qn", "safety_score", "risk_category", "confidence",
        "recommended_preparation",
    ]), rows)


def _sheet_change_impact(wb: Workbook, session: Session, run_id: str) -> None:
    ws = wb.create_sheet("Change Impact")
    # Change impact is computed on demand per component (POST); the report lists whatever
    # has already been persisted for this run.
    from sqlalchemy import select

    from archon.db.models import ChangeImpact, Component

    rows = session.scalars(select(ChangeImpact).where(ChangeImpact.run_id == run_id)).all()
    qn = {
        c.id: c.qualified_name
        for c in session.scalars(
            select(Component).where(Component.id.in_([r.component_id for r in rows]))
        ).all()
    } if rows else {}
    _write_table(
        ws,
        ["component_qn", "direct_dependents", "indirect_dependents", "callers",
         "related_tests", "historical_co_changes", "external_integrations"],
        [{
            "component_qn": qn.get(r.component_id, r.component_id),
            "direct_dependents": len(r.direct_dependents or []),
            "indirect_dependents": len(r.indirect_dependents or []),
            "callers": len(r.callers or []),
            "related_tests": len(r.related_tests or []),
            "historical_co_changes": len(r.historical_co_changes or []),
            "external_integrations": len(r.external_integrations or []),
        } for r in rows],
    )


def _sheet_technical_debt(wb: Workbook, session: Session, run_id: str) -> None:
    ws = wb.create_sheet("Technical Debt")
    rows = _dump(queries.technical_debt(session, run_id))
    _write_table(ws, _headers(rows, [
        "component_qn", "category", "severity", "location", "confidence",
        "recommendation", "impact",
    ]), rows)


def _sheet_test_gaps(wb: Workbook, session: Session, run_id: str) -> None:
    ws = wb.create_sheet("Test Gaps")
    rows = _dump(queries.test_gaps(session, run_id))
    _write_table(ws, _headers(rows, [
        "component_qn", "kind", "priority", "priority_score", "coverage_pct",
        "legacy_risk_score", "change_safety_score",
    ]), rows)


def _sheet_characterization(wb: Workbook, session: Session, run_id: str) -> None:
    ws = wb.create_sheet("Characterization")
    rows = _dump(queries.characterization(session, run_id))
    _write_table(ws, _headers(rows, [
        "component_qn", "baseline_hash", "test_case_id",
    ]), rows)


def _sheet_failures(wb: Workbook, session: Session, run_id: str) -> None:
    ws = wb.create_sheet("Failures")
    rows = _dump(queries.failures(session, run_id))
    _write_table(ws, _headers(rows, [
        "test_identifier", "exception_type", "message", "reproducible", "occurrences",
    ]), rows)


def _sheet_repairs(wb: Workbook, session: Session, run_id: str) -> None:
    ws = wb.create_sheet("Repairs")
    inv = _dump(queries.investigations(session, run_id))
    pat = _dump(queries.patches(session, run_id))
    ver = _dump(queries.verifications(session, run_id))
    ws.append(["Investigations"])
    ws[ws.max_row][0].font = _HEADER_FONT
    _write_table(ws, _headers(inv, ["summary", "confidence", "ai_schema_version"]), inv)
    ws.append([])
    ws.append(["Candidate patches"])
    ws[ws.max_row][0].font = _HEADER_FONT
    _write_table(ws, _headers(pat, ["strategy", "state", "lines_added", "lines_removed", "rank_score"]), pat)
    ws.append([])
    ws.append(["Verifications"])
    ws[ws.max_row][0].font = _HEADER_FONT
    _write_table(ws, _headers(ver, ["patch_id", "verdict", "original_failure_fixed", "regression_pass"]), ver)


def _sheet_modernization(wb: Workbook, session: Session, run_id: str) -> None:
    ws = wb.create_sheet("Modernization")
    rows = _dump(queries.modernization(session, run_id))
    _write_table(ws, _headers(rows, [
        "order_index", "target", "strategy", "risk", "effort", "impact",
        "rationale", "confidence", "classification",
    ]), rows)


def _sheet_software_archaeology(wb: Workbook, session: Session, run_id: str) -> None:
    ws = wb.create_sheet("Software Archaeology")
    evo = queries.evolution(session, run_id)
    if evo is not None:
        _kv(ws, [
            ("Total commits", evo.total_commits),
            ("Analyzed commits", evo.analyzed_commits),
            ("Span (days)", evo.span_days),
            ("Distinct authors", evo.authors),
        ])
        ws.append([])
    ws.append(["Behaviour reconstructions"])
    ws[ws.max_row][0].font = _HEADER_FONT
    beh = _dump(queries.behavior(session, run_id))
    _write_table(ws, _headers(beh, ["component_qn", "purpose", "classification", "confidence"]), beh)
    ws.append([])
    ws.append(["Hidden assumptions"])
    ws[ws.max_row][0].font = _HEADER_FONT
    asm = _dump(queries.assumptions(session, run_id))
    _write_table(ws, _headers(asm, ["component_qn", "kind", "risk", "description", "suggested_test", "confidence"]), asm)


def _sheet_incident_memory(wb: Workbook, session: Session, run_id: str) -> None:
    ws = wb.create_sheet("Incident Memory")
    rows = _dump(queries.incidents_for_run(session, run_id))
    _write_table(ws, _headers(rows, [
        "failure_signature", "failure_summary", "root_cause", "confidence", "patch_id",
    ]), rows)


_SHEETS = (
    _sheet_executive_summary,
    _sheet_repository_understanding,
    _sheet_architecture,
    _sheet_legacy_dna,
    _sheet_change_safety,
    _sheet_change_impact,
    _sheet_technical_debt,
    _sheet_test_gaps,
    _sheet_characterization,
    _sheet_failures,
    _sheet_repairs,
    _sheet_modernization,
    _sheet_software_archaeology,
    _sheet_incident_memory,
)

SHEET_NAMES = (
    "Executive Summary", "Repository Understanding", "Architecture", "Legacy DNA",
    "Change Safety", "Change Impact", "Technical Debt", "Test Gaps", "Characterization",
    "Failures", "Repairs", "Modernization", "Software Archaeology", "Incident Memory",
)


def build_report(session: Session, run_id: str) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet
    for fn in _SHEETS:
        fn(wb, session, run_id)
    return wb


def report_bytes(session: Session, run_id: str) -> bytes:
    buf = BytesIO()
    build_report(session, run_id).save(buf)
    return buf.getvalue()

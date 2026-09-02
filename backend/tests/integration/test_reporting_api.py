"""Phase 13 - Excel reporting API (spec sections 49-50).

Seeds an analysis-scored run directly (no sandbox), then asserts
``GET /runs/{id}/report.xlsx`` is a 14-sheet workbook whose cells match the JSON API.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from archon.core.artifacts import write_json
from archon.db.base import session_scope
from archon.db.models import (
    AnalysisRun,
    ChangeAssessment,
    Component,
    Hotspot,
    LegacyDNA,
    Repository,
    RepositorySnapshot,
    TechnicalDebtFinding,
)
from archon.domain.enums import (
    ComponentKind,
    ProviderKind,
    RunMode,
    RunState,
    Stage,
    TechDebtCategory,
    TechDebtSeverity,
)
from archon.reporting.workbook import SHEET_NAMES


def _seed(url: str = "/tmp/report-api") -> str:
    with session_scope() as s:
        repo = Repository(provider=ProviderKind.LOCAL, url=url, name="rep")
        s.add(repo)
        s.flush()
        snap = RepositorySnapshot(repository_id=repo.id, commit_sha="c" * 40)
        s.add(snap)
        s.flush()
        run = AnalysisRun(
            repository_id=repo.id, snapshot_id=snap.id, mode=RunMode.ANALYSIS_ONLY,
            state=RunState.COMPLETED, last_completed_stage=Stage.ANALYZING_CHANGE_IMPACT,
        )
        s.add(run)
        s.flush()

        comps = {}
        for qn in ("pkg.risky", "pkg.stable"):
            c = Component(
                snapshot_id=snap.id, kind=ComponentKind.MODULE, name=qn.split(".")[-1],
                qualified_name=qn, path=qn.replace(".", "/") + ".py", role="domain",
                metrics={"architecture": {"fan_in": 3, "fan_out": 1, "in_cycle": False}},
            )
            s.add(c)
            s.flush()
            comps[qn] = c

        s.add(LegacyDNA(
            run_id=run.id, snapshot_id=snap.id, component_id=comps["pkg.risky"].id,
            coverage=0.0, coverage_is_proxy=True, complexity=18, churn=40, coupling=9,
            legacy_risk_score=81.0, category="HIGH", confidence=0.8,
            produced_by="legacy_risk.v1", factor_breakdown={},
        ))
        s.add(LegacyDNA(
            run_id=run.id, snapshot_id=snap.id, component_id=comps["pkg.stable"].id,
            coverage=0.9, coverage_is_proxy=True, complexity=2, churn=1, coupling=1,
            legacy_risk_score=8.0, category="LOW", confidence=0.8,
            produced_by="legacy_risk.v1", factor_breakdown={},
        ))
        s.add(Hotspot(
            run_id=run.id, snapshot_id=snap.id, component_id=comps["pkg.risky"].id,
            score=78.0, classification="RISKY", reasons={}, engine_version="hotspot.v1",
        ))
        s.add(ChangeAssessment(
            run_id=run.id, snapshot_id=snap.id, component_id=comps["pkg.risky"].id,
            engine_version="change_safety.v1", safety_score=32.0, risk_category="RISKY",
            recommended_preparation=["add tests"], confidence=0.8, factor_breakdown={},
            produced_by="change_safety.v1",
        ))
        s.add(TechnicalDebtFinding(
            run_id=run.id, snapshot_id=snap.id, component_id=comps["pkg.risky"].id,
            category=TechDebtCategory.BROAD_EXCEPT, location="pkg/risky.py:12",
            severity=TechDebtSeverity.HIGH, confidence=0.8, produced_by="tech_debt.v1",
        ))
        write_json(s, run.id, "understanding", {
            "score": 55.0, "confidence": 0.55,
            "dimensions": {"architecture": 0.6, "testing": 0.3, "behavior": 0.7,
                           "dependency": 0.5, "historical": 0.5, "configuration": 0.5},
            "evidence_coverage": {"modules_total": 2},
        })
        write_json(s, run.id, "architecture_graph", {
            "cycles": [], "layering_violations": [],
            "components": {"nodes": [], "links": []}, "modules": {"nodes": [], "links": []},
        })
        s.flush()
        return run.id


def test_report_xlsx_has_14_sheets_and_matches_api(client):
    run_id = _seed()
    resp = client.get(f"/runs/{run_id}/report.xlsx")
    assert resp.status_code == 200, resp.text
    assert "spreadsheetml.sheet" in resp.headers["content-type"]
    assert 'filename="ARCHON_Legacy_Analysis.xlsx"' in resp.headers["content-disposition"]

    wb = load_workbook(BytesIO(resp.content))
    assert wb.sheetnames == list(SHEET_NAMES)

    # Legacy DNA sheet rows match the JSON endpoint (same component set + scores)
    api_ld = {r["component_qn"]: r for r in client.get(f"/runs/{run_id}/legacy-dna").json()}
    ws = wb["Legacy DNA"]
    header = [c.value for c in ws[1]]
    qn_i = header.index("component_qn")
    score_i = header.index("legacy_risk_score")
    sheet_ld = {ws.cell(r, qn_i + 1).value: ws.cell(r, score_i + 1).value
                for r in range(2, ws.max_row + 1) if ws.cell(r, qn_i + 1).value}
    assert set(sheet_ld) == set(api_ld)
    for qn, score in sheet_ld.items():
        assert abs(score - api_ld[qn]["legacy_risk_score"]) < 1e-6

    # Technical Debt sheet mirrors the API
    api_td = client.get(f"/runs/{run_id}/technical-debt").json()
    ws = wb["Technical Debt"]
    assert ws.max_row - 1 == len(api_td) == 1
    assert ws.cell(2, [c.value for c in ws[1]].index("category") + 1).value == "BROAD_EXCEPT"

    # Executive Summary is a key/value sheet that names the run
    exec_vals = {ws.cell(r, 1).value: ws.cell(r, 2).value
                 for ws in [wb["Executive Summary"]] for r in range(2, ws.max_row + 1)}
    assert exec_vals["Run id"] == run_id
    assert exec_vals["Components scored (Legacy DNA)"] == 2


def test_report_xlsx_404_on_unknown_run(client):
    assert client.get("/runs/nope/report.xlsx").status_code == 404

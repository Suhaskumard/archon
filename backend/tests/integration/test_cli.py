"""`archon` CLI headless ingestion (spec section 20, plan Phase 1)."""

from __future__ import annotations

import json

from typer.testing import CliRunner

from archon.cli.main import app

runner = CliRunner()


def test_analyze_local_repo_end_to_end(test_repo):
    result = runner.invoke(app, ["analyze", str(test_repo), "--wait"])
    assert result.exit_code == 0, result.output
    # the final JSON blob is printed after the "run ... queued" line
    payload = json.loads(result.output[result.output.index("{"):])
    assert payload["state"] == "COMPLETED"
    assert payload["last_completed_stage"] == "SNAPSHOTTING"
    assert payload["snapshot"]["support_level"] == "SUPPORTED"
    assert len(payload["snapshot"]["commit_sha"]) == 40
    assert any(e["classification"] == "FACT" for e in payload["evidence"])


def test_analyze_rejects_bad_target():
    result = runner.invoke(app, ["analyze", "definitely not a repo"])
    assert result.exit_code != 0
    assert "INVALID_REPOSITORY_URL" in result.output or "cannot determine a provider" in str(
        result.exception
    )


def test_db_upgrade_is_idempotent():
    assert runner.invoke(app, ["db-upgrade"]).exit_code == 0
    assert runner.invoke(app, ["db-upgrade"]).exit_code == 0


def test_bulk_import_and_report(tmp_path, test_repo):
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.append(["Repository URL", "Branch", "Analysis Mode", "Priority"])
    wb.active.append([str(test_repo), None, "ANALYSIS_ONLY", 5])
    xlsx = tmp_path / "repos.xlsx"
    wb.save(xlsx)

    r = runner.invoke(app, ["bulk-import", str(xlsx)])
    assert r.exit_code == 0, r.output
    assert "1 run(s) queued" in r.output
    run_id = next(
        line.split("run=")[1].strip()
        for line in r.output.splitlines() if "run=" in line
    )

    out = tmp_path / "report.xlsx"
    r2 = runner.invoke(app, ["report", run_id, "--out", str(out)])
    assert r2.exit_code == 0, r2.output
    assert out.exists() and out.stat().st_size > 0

    from openpyxl import load_workbook

    assert len(load_workbook(out).sheetnames) == 14
